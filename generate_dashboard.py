#!/usr/bin/env python3
"""
Genera un dashboard HTML estatico (docs/index.html) a partir de holded_reviews.xlsx.
Pensado para publicarse con GitHub Pages sirviendo la carpeta docs/.

Uso:
    python generate_dashboard.py
"""

import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

INPUT_XLSX = "holded_reviews.xlsx"
OUTPUT_DIR = Path("docs")
OUTPUT_HTML = OUTPUT_DIR / "index.html"
SHEET_NAME = "Reviews"

VERTICAL_ORDER = [
    "Facturación", "Contabilidad", "Precio", "Soporte", "Conciliación", "Banco",
    "General", "Escáner", "Recursos Humanos", "Nóminas", "CRM", "Proyectos",
    "Reservas", "Inventario", "Fabricación", "Catálogo", "Importación",
    "Analítica", "SII AEAT", "Impuestos", "TPV",
]

VERTICAL_COLORS = {
    "Facturación": "#4F46E5",
    "Contabilidad": "#9333EA",
    "Precio": "#84CC16",
    "Soporte": "#3B82F6",
    "Conciliación": "#0EA5E9",
    "Banco": "#06B6D4",
    "General": "#94A3B8",
    "Escáner": "#8B5CF6",
    "Recursos Humanos": "#EC4899",
    "Nóminas": "#E11D48",
    "CRM": "#F59E0B",
    "Proyectos": "#65A30D",
    "Reservas": "#C026D3",
    "Inventario": "#10B981",
    "Fabricación": "#F97316",
    "Catálogo": "#14B8A6",
    "Importación": "#0D9488",
    "Analítica": "#7C2D12",
    "SII AEAT": "#6366F1",
    "Impuestos": "#EAB308",
    "TPV": "#EF4444",
}

RATING_RE = re.compile(r"^(\d+)/5")


def load_reviews():
    wb = load_workbook(INPUT_XLSX)
    ws = wb[SHEET_NAME]
    reviews = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        fecha, guid, titulo, review, rating, reviewer, pais, vertical, link = row
        m = RATING_RE.match(rating or "")
        rating_num = int(m.group(1)) if m else None
        reviews.append({
            "fecha": str(fecha) if fecha else "",
            "titulo": titulo or "",
            "review": review or "",
            "rating": rating_num,
            "reviewer": reviewer or "",
            "pais": pais or "",
            "vertical": vertical or "General",
            "link": link or "",
        })
    # mas recientes primero
    reviews.sort(key=lambda r: r["fecha"], reverse=True)
    return reviews


def aggregate(reviews):
    by_vertical = defaultdict(list)
    for r in reviews:
        by_vertical[r["vertical"]].append(r)

    stats = {}
    for vertical in VERTICAL_ORDER:
        items = by_vertical.get(vertical, [])
        ratings = [r["rating"] for r in items if r["rating"] is not None]
        avg = round(sum(ratings) / len(ratings), 2) if ratings else None
        dist = {str(n): sum(1 for r in ratings if r == n) for n in range(1, 6)}
        stats[vertical] = {
            "count": len(items),
            "avg_rating": avg,
            "distribution": dist,
        }

    all_ratings = [r["rating"] for r in reviews if r["rating"] is not None]
    overall_avg = round(sum(all_ratings) / len(all_ratings), 2) if all_ratings else None
    negative_count = sum(1 for r in all_ratings if r <= 2)

    return {
        "total": len(reviews),
        "overall_avg": overall_avg,
        "negative_count": negative_count,
        "negative_pct": round(100 * negative_count / len(reviews), 1) if reviews else 0,
        "by_vertical": stats,
    }


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>Dashboard de Reviews · Holded (Trustpilot)</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.5.0/chart.umd.min.js"></script>
<style>
  :root {
    --bg: #F5F6FA;
    --panel: #FFFFFF;
    --text: #1A1F36;
    --muted: #6B7280;
    --border: #E5E7EB;
    --accent: #4F46E5;
    --good: #16A34A;
    --bad: #DC2626;
    --warn: #D97706;
  }
  * { box-sizing: border-box; }
  body {
    margin: 0;
    background: var(--bg);
    color: var(--text);
    font-family: "Inter", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
  }
  header {
    padding: 32px clamp(16px, 4vw, 48px) 20px;
    border-bottom: 1px solid var(--border);
    background: var(--panel);
  }
  header h1 {
    margin: 0 0 4px;
    font-size: 22px;
    font-weight: 700;
    letter-spacing: -0.01em;
  }
  header p {
    margin: 0;
    color: var(--muted);
    font-size: 14px;
  }
  main {
    padding: 24px clamp(16px, 4vw, 48px) 60px;
    max-width: 1280px;
    margin: 0 auto;
  }
  .kpis {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 16px;
    margin-bottom: 24px;
  }
  .kpi {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 18px 20px;
  }
  .kpi .label {
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--muted);
    margin-bottom: 6px;
  }
  .kpi .value {
    font-size: 28px;
    font-weight: 700;
    font-variant-numeric: tabular-nums;
  }
  .kpi .value.bad { color: var(--bad); }
  .kpi .value.good { color: var(--good); }
  .grid-2 {
    display: grid;
    grid-template-columns: 1.2fr 1fr;
    gap: 16px;
    margin-bottom: 16px;
  }
  @media (max-width: 900px) {
    .grid-2 { grid-template-columns: 1fr; }
  }
  .panel {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 20px;
  }
  .panel h2 {
    margin: 0 0 16px;
    font-size: 15px;
    font-weight: 600;
  }
  .chart-wrap { position: relative; height: 320px; }
  .filters {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin: 20px 0 16px;
  }
  .filter-btn {
    border: 1px solid var(--border);
    background: var(--panel);
    color: var(--text);
    padding: 6px 12px;
    border-radius: 999px;
    font-size: 13px;
    cursor: pointer;
    transition: all 0.15s ease;
  }
  .filter-btn:hover { border-color: var(--accent); }
  .filter-btn.active {
    background: var(--accent);
    border-color: var(--accent);
    color: white;
  }
  table {
    width: 100%;
    border-collapse: collapse;
    font-size: 13.5px;
  }
  thead th {
    text-align: left;
    padding: 10px 12px;
    color: var(--muted);
    font-weight: 600;
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    border-bottom: 1px solid var(--border);
  }
  tbody td {
    padding: 12px;
    border-bottom: 1px solid var(--border);
    vertical-align: top;
  }
  tbody tr:hover { background: #FAFAFC; }
  .tag {
    display: inline-block;
    padding: 2px 9px;
    border-radius: 999px;
    font-size: 11.5px;
    font-weight: 600;
    color: white;
    white-space: nowrap;
  }
  .rating-cell { font-variant-numeric: tabular-nums; font-weight: 700; }
  .rating-cell.r1, .rating-cell.r2 { color: var(--bad); }
  .rating-cell.r3 { color: var(--warn); }
  .rating-cell.r4, .rating-cell.r5 { color: var(--good); }
  .review-title { font-weight: 600; margin-bottom: 2px; }
  .review-text { color: var(--muted); max-width: 480px; }
  .meta { color: var(--muted); font-size: 12.5px; }
  .empty-state {
    text-align: center;
    padding: 40px;
    color: var(--muted);
  }
  .review-columns {
    display: flex;
    gap: 14px;
    overflow-x: auto;
    padding-bottom: 4px;
  }
  .review-column {
    flex: 0 0 260px;
    display: flex;
    flex-direction: column;
    gap: 12px;
  }
  .review-column-header {
    display: flex;
    align-items: center;
    gap: 8px;
    font-weight: 700;
    font-size: 13px;
    padding-bottom: 8px;
    border-bottom: 2px solid var(--border);
    margin-bottom: 2px;
  }
  .review-column-header .dot {
    width: 10px;
    height: 10px;
    border-radius: 50%;
    flex: 0 0 auto;
  }
  .review-column-header .count {
    color: var(--muted);
    font-weight: 500;
    margin-left: auto;
  }
  .review-card {
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 16px;
    background: var(--panel);
  }
  .review-card .stars {
    font-size: 15px;
    letter-spacing: 1px;
    color: #D1D5DB;
    margin-bottom: 6px;
  }
  .review-card .stars .filled { color: #F59E0B; }
  .review-card .author {
    font-weight: 600;
    font-size: 13.5px;
    margin-bottom: 2px;
  }
  .review-card .author-meta {
    color: var(--muted);
    font-size: 12px;
    margin-bottom: 10px;
  }
  .review-card .card-title {
    font-weight: 600;
    font-size: 13.5px;
    margin-bottom: 4px;
  }
  .review-card .card-text {
    color: var(--muted);
    font-size: 13px;
    line-height: 1.45;
  }
  .read-more-btn {
    display: block;
    margin-top: 6px;
    background: none;
    border: none;
    padding: 0;
    color: var(--accent);
    font-size: 12px;
    font-weight: 600;
    cursor: pointer;
  }
  .read-more-btn:hover { text-decoration: underline; }
  footer {
    text-align: center;
    color: var(--muted);
    font-size: 12px;
    padding: 20px;
  }
</style>
</head>
<body>
<header>
  <h1>Dashboard de Reviews — Holded en Trustpilot</h1>
  <p>Actualizado automaticamente desde el feed RSS · Generado el __GENERATED_AT__</p>
</header>
<main>
  <div class="kpis">
    <div class="kpi">
      <div class="label">Total reviews</div>
      <div class="value">__TOTAL__</div>
    </div>
    <div class="kpi">
      <div class="label">Rating medio</div>
      <div class="value">__OVERALL_AVG__ / 5</div>
    </div>
    <div class="kpi">
      <div class="label">Reviews negativas (1-2 estrellas)</div>
      <div class="value bad">__NEGATIVE_COUNT__ (__NEGATIVE_PCT__%)</div>
    </div>
    <div class="kpi">
      <div class="label">Verticales con reviews</div>
      <div class="value">__VERTICALS_WITH_DATA__</div>
    </div>
  </div>

  <div class="grid-2">
    <div class="panel">
      <h2>Volumen de reviews por vertical</h2>
      <div class="chart-wrap"><canvas id="chartVolumen"></canvas></div>
    </div>
    <div class="panel">
      <h2>Rating medio por vertical</h2>
      <div class="chart-wrap"><canvas id="chartRating"></canvas></div>
    </div>
  </div>

  <div class="panel">
    <h2>Últimas reviews por vertical</h2>
    <div class="review-columns" id="reviewCards"></div>
  </div>

  <div class="panel">
    <h2>Reviews</h2>
    <div class="filters" id="filters"></div>
    <div class="filters" id="ratingFilters"></div>
    <div id="tableWrap">
      <table>
        <thead>
          <tr>
            <th>Fecha</th>
            <th>Vertical</th>
            <th>Rating</th>
            <th>Review</th>
            <th>Reviewer</th>
          </tr>
        </thead>
        <tbody id="tableBody"></tbody>
      </table>
      <div class="empty-state" id="emptyState" style="display:none;">Sin reviews en este vertical todavia.</div>
    </div>
  </div>
</main>
<footer>Datos obtenidos de Trustpilot via RSS · Solo uso interno</footer>

<script>
const REVIEWS = __REVIEWS_JSON__;
const STATS = __STATS_JSON__;
const COLORS = __COLORS_JSON__;
const VERTICAL_ORDER = __ORDER_JSON__;

const verticalsWithData = VERTICAL_ORDER.filter(v => STATS.by_vertical[v].count > 0);

// --- Charts ---
// Envuelto en try/catch: si Chart.js no carga (CDN caido, bloqueado, etc.)
// el resto del dashboard (tarjetas, filtros, tabla) debe seguir funcionando.
try {
  new Chart(document.getElementById('chartVolumen'), {
    type: 'bar',
    data: {
      labels: verticalsWithData,
      datasets: [{
        data: verticalsWithData.map(v => STATS.by_vertical[v].count),
        backgroundColor: verticalsWithData.map(v => COLORS[v]),
        borderRadius: 6,
      }]
    },
    options: {
      responsive: true,
      maintainAspectRatio: false,
      plugins: { legend: { display: false } },
      scales: {
        y: { beginAtZero: true, ticks: { precision: 0 } }
      }
    }
  });

  new Chart(document.getElementById('chartRating'), {
    type: 'bar',
    data: {
      labels: verticalsWithData,
      datasets: [{
        data: verticalsWithData.map(v => STATS.by_vertical[v].avg_rating ?? 0),
        backgroundColor: verticalsWithData.map(v => COLORS[v]),
        borderRadius: 6,
      }]
    },
    options: {
      responsive: true,
      maintainAspectRatio: false,
      plugins: { legend: { display: false } },
      scales: {
        y: { beginAtZero: true, max: 5 }
      }
    }
  });
} catch (err) {
  console.error('No se pudieron renderizar los graficos:', err);
}

// --- Ultimas reviews por vertical (columnas con puntuacion y autor) ---
const CARDS_PER_COLUMN = 6;

function renderStars(rating) {
  const r = rating || 0;
  let html = '';
  for (let i = 1; i <= 5; i++) {
    html += `<span class="${i <= r ? 'filled' : ''}">★</span>`;
  }
  return html;
}

const CARD_TEXTS = [];
let cardTextSeq = 0;

function renderReviewCard(r) {
  const isLong = r.review.length > 160;
  const reviewShort = isLong ? r.review.slice(0, 160) + '…' : r.review;
  const idx = cardTextSeq++;
  CARD_TEXTS[idx] = r.review;
  return `
    <div class="review-card">
      <div class="stars">${renderStars(r.rating)}</div>
      <div class="author">${escapeHtml(r.reviewer || 'Anonimo')}</div>
      <div class="author-meta">${escapeHtml(r.pais || '')}${r.pais ? ' · ' : ''}${escapeHtml(r.fecha)}</div>
      <div class="card-title">${escapeHtml(r.titulo)}</div>
      <div class="card-text" data-idx="${idx}">${escapeHtml(reviewShort)}</div>
      ${isLong ? `<button class="read-more-btn" data-target="card" data-idx="${idx}">Leer más</button>` : ''}
    </div>
  `;
}

function renderCards() {
  const wrap = document.getElementById('reviewCards');

  if (REVIEWS.length === 0) {
    wrap.innerHTML = '<div class="empty-state">Sin reviews todavia.</div>';
    return;
  }

  cardTextSeq = 0;
  wrap.innerHTML = verticalsWithData.map(vertical => {
    const items = REVIEWS.filter(r => r.vertical === vertical).slice(0, CARDS_PER_COLUMN);
    const color = COLORS[vertical] || '#94A3B8';
    return `
      <div class="review-column">
        <div class="review-column-header">
          <span class="dot" style="background:${color}"></span>
          <span>${escapeHtml(vertical)}</span>
          <span class="count">${STATS.by_vertical[vertical].count}</span>
        </div>
        ${items.map(renderReviewCard).join('')}
      </div>
    `;
  }).join('');
}

// --- Filtros + tabla ---
let activeFilter = 'Todos';
let activeRating = 'Todos';

function renderFilters() {
  const filtersEl = document.getElementById('filters');
  const options = ['Todos', ...verticalsWithData];
  filtersEl.innerHTML = options.map(v => {
    const count = v === 'Todos' ? REVIEWS.length : STATS.by_vertical[v].count;
    return `<button class="filter-btn ${v === activeFilter ? 'active' : ''}" data-v="${v}">${v} (${count})</button>`;
  }).join('');
  filtersEl.querySelectorAll('.filter-btn').forEach(btn => {
    btn.addEventListener('click', () => {
      activeFilter = btn.dataset.v;
      renderFilters();
      renderTable();
    });
  });
}

function renderRatingFilters() {
  const filtersEl = document.getElementById('ratingFilters');
  const ratings = [5, 4, 3, 2, 1];
  const countFor = (n) => REVIEWS.filter(r => r.rating === n).length;
  const options = [
    { v: 'Todos', label: 'Todos los ratings', count: REVIEWS.length },
    ...ratings.map(n => ({ v: String(n), label: '★'.repeat(n), count: countFor(n) })),
  ];
  filtersEl.innerHTML = options.map(o => {
    return `<button class="filter-btn ${o.v === activeRating ? 'active' : ''}" data-r="${o.v}">${o.label} (${o.count})</button>`;
  }).join('');
  filtersEl.querySelectorAll('.filter-btn').forEach(btn => {
    btn.addEventListener('click', () => {
      activeRating = btn.dataset.r;
      renderRatingFilters();
      renderTable();
    });
  });
}

function escapeHtml(str) {
  const div = document.createElement('div');
  div.textContent = str;
  return div.innerHTML;
}

const ROW_TEXTS = [];

function renderTable() {
  const tbody = document.getElementById('tableBody');
  const emptyState = document.getElementById('emptyState');
  let rows = activeFilter === 'Todos' ? REVIEWS : REVIEWS.filter(r => r.vertical === activeFilter);
  if (activeRating !== 'Todos') {
    rows = rows.filter(r => r.rating === Number(activeRating));
  }

  if (rows.length === 0) {
    tbody.innerHTML = '';
    emptyState.style.display = 'block';
    return;
  }
  emptyState.style.display = 'none';

  ROW_TEXTS.length = 0;
  tbody.innerHTML = rows.map((r, idx) => {
    const color = COLORS[r.vertical] || '#94A3B8';
    const ratingClass = r.rating ? `r${r.rating}` : '';
    const isLong = r.review.length > 220;
    const reviewShort = isLong ? r.review.slice(0, 220) + '…' : r.review;
    ROW_TEXTS[idx] = r.review;
    return `
      <tr>
        <td class="meta">${escapeHtml(r.fecha)}</td>
        <td><span class="tag" style="background:${color}">${escapeHtml(r.vertical)}</span></td>
        <td class="rating-cell ${ratingClass}">${r.rating ? r.rating + '/5' : '—'}</td>
        <td>
          <div class="review-title">${escapeHtml(r.titulo)}</div>
          <div class="review-text" data-idx="${idx}">${escapeHtml(reviewShort)}</div>
          ${isLong ? `<button class="read-more-btn" data-target="row" data-idx="${idx}">Leer más</button>` : ''}
        </td>
        <td class="meta">${escapeHtml(r.reviewer)}${r.pais ? ', ' + escapeHtml(r.pais) : ''}</td>
      </tr>
    `;
  }).join('');
}

function truncate(text, limit) {
  return text.length > limit ? text.slice(0, limit) + '…' : text;
}

function handleReadMoreClick(e) {
  const btn = e.target.closest('.read-more-btn');
  if (!btn) return;
  const idx = btn.dataset.idx;
  const isCard = btn.dataset.target === 'card';
  const store = isCard ? CARD_TEXTS : ROW_TEXTS;
  const limit = isCard ? 160 : 220;
  const selector = `${isCard ? '.card-text' : '.review-text'}[data-idx="${idx}"]`;
  const textEl = btn.parentElement.querySelector(selector) || btn.previousElementSibling;
  const fullText = store[idx];
  const expanded = btn.dataset.expanded === 'true';
  textEl.textContent = expanded ? truncate(fullText, limit) : fullText;
  btn.textContent = expanded ? 'Leer más' : 'Leer menos';
  btn.dataset.expanded = expanded ? 'false' : 'true';
}

document.getElementById('reviewCards').addEventListener('click', handleReadMoreClick);
document.getElementById('tableWrap').addEventListener('click', handleReadMoreClick);

renderCards();
renderFilters();
renderRatingFilters();
renderTable();
</script>
</body>
</html>
"""


def render_html(reviews, stats):
    html = HTML_TEMPLATE
    html = html.replace("__GENERATED_AT__", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
    html = html.replace("__TOTAL__", str(stats["total"]))
    html = html.replace("__OVERALL_AVG__", str(stats["overall_avg"] if stats["overall_avg"] is not None else "—"))
    html = html.replace("__NEGATIVE_COUNT__", str(stats["negative_count"]))
    html = html.replace("__NEGATIVE_PCT__", str(stats["negative_pct"]))
    verticals_with_data = sum(1 for v in VERTICAL_ORDER if stats["by_vertical"][v]["count"] > 0)
    html = html.replace("__VERTICALS_WITH_DATA__", str(verticals_with_data))
    html = html.replace("__REVIEWS_JSON__", json.dumps(reviews, ensure_ascii=False))
    html = html.replace("__STATS_JSON__", json.dumps(stats, ensure_ascii=False))
    html = html.replace("__COLORS_JSON__", json.dumps(VERTICAL_COLORS, ensure_ascii=False))
    html = html.replace("__ORDER_JSON__", json.dumps(VERTICAL_ORDER, ensure_ascii=False))
    return html


def main():
    reviews = load_reviews()
    stats = aggregate(reviews)

    OUTPUT_DIR.mkdir(exist_ok=True)
    html = render_html(reviews, stats)
    OUTPUT_HTML.write_text(html, encoding="utf-8")

    print(f"Dashboard generado: {OUTPUT_HTML.resolve()}")
    print(f"Total reviews: {stats['total']} | Rating medio: {stats['overall_avg']}")


if __name__ == "__main__":
    sys.exit(main())
