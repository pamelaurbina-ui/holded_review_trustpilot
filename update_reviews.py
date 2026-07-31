#!/usr/bin/env python3
"""
Actualiza un Excel con las reviews de Trustpilot (Holded) usando la API oficial
de Trustpilot Business.

Uso:
    TRUSTPILOT_API_KEY=xxx python update_reviews.py

Comportamiento:
- Pagina la API de reviews (mas recientes primero) hasta encontrar una review
  ya conocida (dedupe por GUID = id de Trustpilot).
- Si el Excel de salida ya existe, añade solo las reviews nuevas.
- Si no existe, lo crea con cabeceras y formato.
"""

import json
import os
import re
import sys
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

API_BASE = "https://api.trustpilot.com/v1"
BUSINESS_UNIT_ID = "5b924bf7477e7d0001af5bc7"
PER_PAGE = 100
MAX_PAGES = 30

OUTPUT_XLSX = "holded_reviews.xlsx"
SHEET_NAME = "Reviews"

HEADERS = ["Fecha", "GUID", "Titulo", "Review", "Rating", "Reviewer", "Pais", "Vertical", "Link"]

# Orden de prioridad: si una review menciona varias palabras clave de distintos
# verticales, gana el primero de esta lista que aparezca (los mas especificos
# primero, "General" siempre al final como fallback).
VERTICAL_KEYWORDS = [
    ("SII AEAT", [
        "sii", "aeat", "hacienda", "modelo 303", "modelo 130", "modelo 111",
        "suministro inmediato", "agencia tributaria", "tax authority", "tax agency",
    ]),
    ("Impuestos", [
        "impuesto", "impuestos", "iva", "vat", "tax rate", "tipo de iva",
        "declaracion de impuestos", "declaración de impuestos", "tax return",
    ]),
    ("TPV", [
        "tpv", "pos", "punto de venta", "point of sale", "caja registradora",
        "terminal de venta", "datafono", "datáfono",
    ]),
    ("Facturación", [
        "factura", "invoic", "billing", "presupuesto", "quote", "cobro",
        "recurring invoice", "facturacion recurrente",
    ]),
    ("Contabilidad", [
        "contabilidad", "accounting", "asiento contable", "libro mayor",
        "plan contable", "cuenta de resultados", "balance contable", "contable",
    ]),
    ("Precio", [
        "precio", "precios", "caro", "carísimo", "carisimo", "price", "pricing",
        "expensive", "overpriced", "cost", "costly", "value for money",
        "relacion calidad-precio", "relación calidad-precio", "subida de precio",
        "price increase", "aumento de precio",
    ]),
    ("Soporte", [
        "soporte", "support team", "customer support", "customer service",
        "atencion al cliente", "atención al cliente", "servicio al cliente",
        "asistencia tecnica", "asistencia técnica", "help desk", "helpdesk",
        "no responden", "tardan en responder", "sin respuesta", "unresponsive",
        "chat de soporte", "equipo de soporte",
    ]),
    ("Conciliación", [
        "concilia", "reconcil", "extracto bancario", "bank statement",
        "sincroniza con el banco", "bank sync", "sync my bank", "sync bank",
        "movimientos bancarios",
    ]),
    ("Banco", [
        "banco", "bancaria", "bancario", "bank account", "cuenta bancaria",
        "conectar banco", "vincular banco", "banking connection",
        "conexion bancaria", "conexión bancaria",
    ]),
    ("Tesorería", [
        "tesoreria", "tesorería", "treasury", "flujo de caja", "cash flow",
        "gestion de caja", "gestión de caja", "liquidez", "cash management",
        "prevision de tesoreria", "previsión de tesorería",
    ]),
    ("Wallet", [
        "wallet", "monedero", "e-wallet", "billetera", "cartera digital",
        "monedero virtual",
    ]),
    ("Nóminas", [
        "nomina", "nóminas", "payroll", "salario", "sueldo",
    ]),
    ("Recursos Humanos", [
        "rrhh", "hr", "human resources", "empleado", "employee",
        "vacaciones", "ausencias", "fichaje",
    ]),
    ("CRM", [
        "crm", "lead", "pipeline", "embudo de ventas", "sales funnel",
        "gestion de clientes", "gestión de clientes", "oportunidad de venta",
    ]),
    ("Proyectos", [
        "proyecto", "proyectos", "project management", "gestion de proyectos",
        "gestión de proyectos", "task management", "gestion de tareas",
        "gestión de tareas",
    ]),
    ("Reservas", [
        "reserva", "reservas", "booking", "sistema de reservas",
        "cita previa", "agenda de citas", "appointment",
    ]),
    ("Inventario", [
        "inventar", "inventory", "stock", "almacen", "almacén", "warehouse",
        "existencias",
    ]),
    ("Fabricación", [
        "fabrica", "manufactur", "produccion", "producción", "production",
        "escandallo", "bom", "bill of materials", "orden de produccion",
    ]),
    ("Catálogo", [
        "catalog", "catálog", "ficha de producto", "product listing",
        "variantes de producto",
    ]),
    ("Importación", [
        "importacion", "importación", "importar", "import data",
        "csv import", "migracion de datos", "migración de datos",
    ]),
    ("Analítica", [
        "analitica", "analítica", "analytics", "informes", "reportes",
        "reporting", "estadisticas", "estadísticas", "kpi",
    ]),
    ("Escáner", [
        "escaner", "escáner", "scanner", "escanear", "ocr", "scan receipt",
        "escanea tickets", "escanea facturas",
    ]),
    ("ERP", [
        "erp", "sistema erp", "software erp", "gestion erp", "gestión erp",
    ]),
]


# Palabras cortas/ambiguas: deben coincidir como palabra EXACTA (limite al
# principio Y al final), para no dar falsos positivos (ej. "ocr" dentro de
# "mediocre", "pos" dentro de "suppose", "hr" dentro de otra palabra).
EXACT_WORD_KEYWORDS = {
    "sii", "pos", "hr", "ocr", "crm", "lead", "bom", "precio", "precios",
    "iva", "kpi", "erp",
}


def classify_vertical(text: str) -> str:
    """Clasifica el texto de una review en un vertical segun palabras clave.
    Para keywords "raiz" (ej. 'factura') solo exige limite de palabra al
    inicio, para que tambien capture 'facturacion' o 'facturation'. Para
    keywords cortas/ambiguas exige palabra exacta completa.
    Devuelve 'General' si no hay coincidencia."""
    lowered = text.lower()
    for vertical, keywords in VERTICAL_KEYWORDS:
        for kw in keywords:
            kw_clean = kw.strip()
            if kw_clean in EXACT_WORD_KEYWORDS:
                pattern = r"\b" + re.escape(kw_clean) + r"\b"
            else:
                pattern = r"\b" + re.escape(kw_clean)
            if re.search(pattern, lowered):
                return vertical
    return "General"

def api_key() -> str:
    key = os.environ.get("TRUSTPILOT_API_KEY")
    if not key:
        raise RuntimeError(
            "Falta la variable de entorno TRUSTPILOT_API_KEY con la API key de Trustpilot Business."
        )
    return key


def fetch_reviews_page(page: int) -> dict:
    params = {
        "apikey": api_key(),
        "perPage": PER_PAGE,
        "page": page,
        "orderBy": "createdat.desc",
    }
    url = f"{API_BASE}/business-units/{BUSINESS_UNIT_ID}/reviews?{urllib.parse.urlencode(params)}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read())


def parse_review(raw: dict) -> dict:
    review_id = raw.get("id", "")
    created_at = raw.get("createdAt") or ""
    fecha = created_at.split("T")[0] if created_at else ""
    stars = raw.get("stars")
    rating = f"{stars}/5" if stars is not None else ""
    consumer = raw.get("consumer") or {}
    titulo = (raw.get("title") or "").strip()
    review_text = (raw.get("text") or "").strip()

    return {
        "fecha": fecha,
        "guid": review_id,
        "titulo": titulo,
        "review": review_text,
        "rating": rating,
        "reviewer": consumer.get("displayName") or "",
        "pais": consumer.get("displayLocation") or "",
        "vertical": classify_vertical(f"{titulo} {review_text}"),
        "link": f"https://www.trustpilot.com/reviews/{review_id}",
    }


def fetch_new_reviews(known_guids: set) -> list:
    """Recorre las reviews mas recientes primero (via API) y se detiene en
    cuanto encuentra una que ya conocemos, asumiendo que todas las anteriores
    a esa (mas antiguas) ya estan importadas."""
    collected = []
    for page in range(1, MAX_PAGES + 1):
        data = fetch_reviews_page(page)
        raw_reviews = data.get("reviews", [])
        if not raw_reviews:
            break
        hit_known = False
        for raw in raw_reviews:
            if raw.get("id", "") in known_guids:
                hit_known = True
                break
            collected.append(parse_review(raw))
        if hit_known:
            break
    return collected


def load_or_create_workbook(path: Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(name="Arial", bold=True)
        cell.alignment = Alignment(horizontal="center")
    return wb, ws


def existing_guids(ws) -> set:
    guids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > 1 and row[1]:
            guids.add(row[1])
    return guids


def autosize_columns(ws):
    widths = {1: 12, 2: 34, 3: 30, 4: 60, 5: 16, 6: 22, 7: 8, 8: 18, 9: 45}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main():
    output_path = Path(OUTPUT_XLSX)

    wb, ws = load_or_create_workbook(output_path)
    known_guids = existing_guids(ws)

    print("Consultando la API de Trustpilot...")
    reviews = fetch_new_reviews(known_guids)
    print(f"Reviews nuevas encontradas: {len(reviews)}")

    new_count = 0
    for r in reviews:
        ws.append([
            r["fecha"], r["guid"], r["titulo"], r["review"],
            r["rating"], r["reviewer"], r["pais"], r["vertical"], r["link"],
        ])
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=ws.max_row, column=col_idx).font = Font(name="Arial")
            ws.cell(row=ws.max_row, column=col_idx).alignment = Alignment(
                vertical="top", wrap_text=(col_idx == 4)
            )
        # Colorear la celda de rating segun sea buena/mala, para lectura rapida
        rating_cell = ws.cell(row=ws.max_row, column=5)
        if r["rating"].startswith(("1", "2")):
            rating_cell.font = Font(name="Arial", color="C0392B", bold=True)
        elif r["rating"].startswith(("4", "5")):
            rating_cell.font = Font(name="Arial", color="1E8449", bold=True)
        new_count += 1

    autosize_columns(ws)
    wb.save(output_path)

    print(f"Reviews nuevas anadidas: {new_count}")
    print(f"Total de reviews en el Excel: {ws.max_row - 1}")
    print(f"Guardado en: {output_path.resolve()}")


if __name__ == "__main__":
    sys.exit(main())
